import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Alignment
import my_person_styles as mps
import df_tools as dftools
import pretty_prints as pp

def xml_2003_reader(xml_file):
    namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    tree = ET.parse(xml_file)
    root = tree.getroot()
    data_rows = []

    for row in root.findall('.//ss:Worksheet/ss:Table/ss:Row', namespaces):
        row_data = []
        for cell in row.findall('ss:Cell', namespaces):
            data = cell.find('ss:Data', namespaces)
            row_data.append(data.text.strip() if data is not None and data.text else "")
        data_rows.append(row_data)

    data_rows.pop()

    column_names = data_rows[0] if data_rows else []
    df_mgt_xml = pd.DataFrame(data_rows[1:], columns=column_names)
    return df_mgt_xml


def mgt_reader(xml_path, workbook_path, wb_sheet_name, export_path):

    df = xml_2003_reader(xml_path)
    wb = pd.read_excel(workbook_path, sheet_name=wb_sheet_name)
    main_df = pd.read_excel(workbook_path, sheet_name='mgt')

    mgt_float_columns = ['Frete', 'Desconto', 'Total da Venda']

    df['Firstname'] = df['Firstname'] + ' ' + df['Lastname']

    new_df = df[[
    'Pedido #', 'Comprado Em', 'Firstname', 'Email', 'Número CPF/CNPJ',
    'Shipping Telephone', 'Frete', 'Desconto', 'Total da Venda',
    'Payment Type', 'Status'
    ]]

    dftools.colum_to_br_currency(new_df, mgt_float_columns)

    new_df = new_df.rename(columns={
        'Pedido #': 'OC',
        'Comprado Em': 'Data',
        'Shipping Telephone': 'Telefone',
        'Firstname': 'Nome',
        'Número CPF/CNPJ': 'CPF',
        'Payment Type': 'Método Pagamento',
        })

    new_df['Data'] = new_df['Data'].apply(dftools.column_to_date)
    new_df['OC'] = pd.to_numeric(new_df['OC'], errors='coerce')

    final_df = dftools.update_df(main_df, new_df, on_column='OC')
    final_df = final_df.sort_values(by='Data')


    with pd.ExcelWriter(
        workbook_path,
        engine='openpyxl',
        mode='a',
        if_sheet_exists='overlay'
        ) as writer:
        final_df.to_excel(writer, sheet_name='mgt', index=False)


    # Estilização
    wb = load_workbook(workbook_path)
    ws = wb['mgt']

    # Verificar estilos do WB e adicionar
    my_named_styles = [mps.br_currency, mps.date]
    mps.styles_verify(my_named_styles, wb)

    date_columns = ['B']
    currencies_columns = ['G', 'H' , 'I']

    style_print_counter = 0
    for row in ws.iter_rows():
        # pp.sleep_with_cls(0.5)
        if style_print_counter == 3:
            style_print_counter = 0
        style_print_counter += 1
        print('Estilizando' + '.' * style_print_counter)

        if any(cell.value is not None for cell in row):
            for cell in row:
                if cell.column_letter in date_columns:
                    cell.style = mps.date.name

                if cell.column_letter in currencies_columns:
                    cell.style = mps.br_currency.name

                cell.border = mps.black_border
                cell.alignment = mps.center_alignment
                cell.font = "Lexend"


    wb.save(workbook_path)
    
    # pp.print_with_cls('Estilização concluída!')

    df_export_path = export_path
    new_df.to_excel(df_export_path, index=False)
