import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Alignment

def magento_reader(magento_xml):
    namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    tree = ET.parse(magento_xml)
    root = tree.getroot()
    data_rows = []

    for row in root.findall('.//ss:Worksheet/ss:Table/ss:Row', namespaces):
        row_data = []
        for cell in row.findall('ss:Cell', namespaces):
            data = cell.find('ss:Data', namespaces)
            row_data.append(data.text.strip() if data is not None and data.text else "")
        data_rows.append(row_data)

    data_rows.pop()

    column_names = data_rows[0] if data_rows else []
    df_magento_xml = pd.DataFrame(data_rows[1:], columns=column_names)
    return df_magento_xml

def date_column_formater(column):
    return pd.to_datetime(column, errors='coerce', dayfirst=True)

def float_column_formater(df, coluns):
    for column in coluns:
        df[column] = df[column].str.replace('R$', '', regex=False)
        df[column] = df[column].str.replace('.', '', regex=False)
        df[column] = df[column].str.replace(',', '.', regex=False)
        df[column] = df[column].astype(float)


magento_file_name = '08-agosto'
magento_path = f'C:\\Users\\jonat\\Desktop\\sl_codes\\{magento_file_name}' + '.xml'
df = magento_reader(magento_path)


new_df = df[[
    'Pedido #', 'Comprado Em', 'Firstname', 'Email', 'Número CPF/CNPJ',
    'Shipping Telephone', 'Frete', 'Desconto', 'Total da Venda',
    'Payment Type', 'Status'
    ]]


magento_float_coluns = ['Frete', 'Desconto', 'Total da Venda']
float_column_formater(new_df, magento_float_coluns)

new_df = new_df.rename(columns={'Pedido #': 'OC'})

new_df['Comprado Em'] = new_df['Comprado Em'].apply(lambda x: x[:10])
new_df['Comprado Em'] = new_df['Comprado Em'].apply(date_column_formater)
new_df['OC'] = pd.to_numeric(new_df['OC'], errors='coerce')


# Estilização dos dados na planilha site
site_sheet_path = 'C:\\Users\\jonat\\Desktop\\sl_codes\\site.xlsx'
wb = load_workbook(site_sheet_path)
ws = wb['Magento']
last_row = ws.max_row + 1
 
# Estilização das células
center_alignment = Alignment(horizontal='center')

border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)


# Named Styles
date_style = NamedStyle(
    name='date_style', 
    number_format="DD/MM/YYYY", 
    )

br_currency_style = NamedStyle(
    name='br_currency_style',
    number_format='R$0.00',
)

my_named_styles = [date_style, br_currency_style]

for style in my_named_styles:
    if style.name not in wb.style_names:
        wb.add_named_style(style)


date_columns = [2]
currencies_columns = [7, 8 , 9]

for r_index, row in new_df.iterrows():
    for c_index, value in enumerate(row, 1):

        cell = ws.cell(row=last_row + r_index, column=c_index, value=value)
        
        if c_index in currencies_columns:
            cell.style = 'br_currency_style'
        
        elif c_index in date_columns:
            cell.style = 'date_style'

        cell.border = border
        cell.alignment=center_alignment
            
wb.save(site_sheet_path)

print('Adição e estilização de dados em site.xlsx "magento" concluída!')

mgt_export_name = f'{magento_file_name} teste'
mgt_export_path = f'C:\\Users\\jonat\\Desktop\\sl_codes\\export{mgt_export_name}' + '.xlsx'
new_df.to_excel(mgt_export_path, index=False)
