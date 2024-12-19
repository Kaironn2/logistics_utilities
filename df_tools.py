import pandas as pd
from openpyxl import load_workbook, Workbook

def update_df(df_main, df_new, on_column):
    existing_id = df_main[on_column].isin(df_new[on_column])

    df_updated = pd.merge(
        df_main[~existing_id],
        df_new,
        on=on_column,
        how='outer',
        suffixes=('', '_new')
    )

    for column in df_new.columns:
        if column != on_column:  # Evita sobreposição do ID
            df_updated[column] = df_updated[f'{column}_new'].combine_first(df_updated[column])
            df_updated.drop(columns=[f'{column}_new'], inplace=True)
    
    return df_updated

def column_to_date(column):
    return pd.to_datetime(str(column)[:10], errors='coerce', dayfirst=True)

def colum_to_br_currency(df, columns):
    for column in columns:
        df[column] = df[column].str.replace('R$', '', regex=False)
        df[column] = df[column].str.replace('.', '', regex=False)
        df[column] = df[column].str.replace(',', '.', regex=False)
        df[column] = df[column].astype(float)

def create_wb_sheet(wb_path, sheets_dict):

    try:
        wb = load_workbook(wb_path)
    except FileNotFoundError:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    for sheet_name, columns_list in sheets_dict.items():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            sheet.append(columns_list)

    wb.save(wb_path)