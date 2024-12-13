import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import df_replacers as dfrep
import logistics_utilities.paths as mpaths
import df_tools as dftools
import re
import my_person_styles as mps

file_name = 'Eccosys HOJE'
export_path = mpaths.sl_codes_path.joinpath(f'{file_name}.xlsx')

def extrair_prazo(texto):
    if isinstance(texto, str):
        padrao = re.compile(r'média\s+(\d+)', re.IGNORECASE)
        match = padrao.search(texto)
        if match:
            return int(match.group(1))
    return None


import_path = mpaths.ecs_path

df = pd.read_csv(import_path, sep=';')

new_df = df[[
    'Numero da Ordem','Número do Pedido', 'Data', 'Data do Pagamento', 'Nome do contato', 'Cód. Rastreamento', 'CEP de entrega', 
    'Cidade do contato', 'UF do Contato', 'Transportadora', 'Forma Frete', 'Observação Interna', 'Data de Entrega'
    ]].copy()

new_df['CEP de entrega'] = new_df['CEP de entrega'].str.replace('.', '', regex=False)
new_df['Prazo Dias Úteis'] = new_df['Observação Interna'].apply(extrair_prazo)
new_df['CNPJ'] = new_df['Observação Interna'].apply(lambda x: dfrep.df_value_replacer(x, dfrep.company_list))
new_df['Transportadora'] = new_df['Transportadora'].apply(lambda x: dfrep.df_value_replacer(x, dfrep.carrier_list))
new_df['Data'] = new_df['Data'].apply(dftools.column_to_date)
new_df['Data do Pagamento'] = new_df['Data do Pagamento'].apply(dftools.column_to_date)
new_df['Data de Entrega'] = new_df['Data de Entrega'].apply(dftools.column_to_date)
# adicionar extração do cupom localizado em Observação Interna e, após isso, apagar a coluna de observação

new_df = new_df[new_df['Nome do contato'] != 'Jose Lindemberg Filho']
new_df = new_df.drop_duplicates(subset='Numero da Ordem', keep='first')
new_df = new_df.sort_values(by='Data', ascending=True)

new_df.to_excel(excel_writer=export_path, index=False)


wb = load_workbook(export_path)
ws = wb.active

def date_stylish(columns_to_format):
    for column in columns_to_format:
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=column)
            cell.style = mps.date

date_columns = [3, 4, 13]
date_stylish(date_columns)

wb.save(export_path)

print(f'Arquivo exportado para: -> {export_path}')